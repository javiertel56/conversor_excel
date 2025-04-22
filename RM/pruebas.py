# =============================
# Importación de librerías
# =============================
import pandas as pd
import numpy as np
import tkinter as tk
from tkinter import filedialog, messagebox
import os
import subprocess
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter

# =============================
# Función de transformación principal
# =============================
def transformar_excel(ruta_entrada, ruta_salida):
    # Leer el archivo Excel original
    df = pd.read_excel(ruta_entrada)

    # Filtrar valores negativos en la columna 'Importe'
    df = df[~df['Importe'].astype(str).str.contains('-', na=False)]
    df = df[df["Importe"] >= 0]

    # Crear columnas auxiliares en minúsculas para facilitar la búsqueda de palabras clave
    df['linea'] = df['Líneas de factura'].fillna('').astype(str).str.lower()
    df['partner'] = df['Partner'].fillna('').astype(str).str.lower()

    # Mapeo de palabras clave a categorías específicas
    keywords_map = {
        'Facta': ['facta'],
        'Master': ['max'],
        'Almacen': ['alm', 'almacen'],
        'Comision': ['comision'],
        'Submarcell': ['submr'],
        'Linea 9': ['linea 9', 'línea 9'],
        'Caja de cobro': ['caja de cobro'],
        'Traspaso': ['traspaso'],
        'SAT': ['sat', 'servicio de administracion', 'servicios de administracion']
    }

    # Función para clasificar un grupo de filas del mismo asiento contable
    def clasificar_asiento(grupo):
        resultado = {}
        resultado['Día'] = pd.to_datetime(grupo['Fecha'].iloc[0]).date()
        resultado['Asiento contable'] = grupo['Asiento contable'].iloc[0]
        resultado['Concepto / Referencia'] = grupo['Líneas de factura'].iloc[0]
        resultado['cargo'] = grupo['Líneas de factura/Débito'].sum()
        resultado['Abono'] = grupo['Líneas de factura/Crédito'].sum()
        resultado['Referencia'] = grupo['Referencia'].iloc[0]

        # Clasificación según palabras clave
        for col, keys in keywords_map.items():
            mask = grupo.apply(lambda row: any(kw in row['linea'] for kw in keys) or any(kw in row['partner'] for kw in keys), axis=1)
            resultado[col] = grupo.loc[mask, 'Líneas de factura/Débito'].sum()

        # Calcular "Otros" y redondeo
        clas_cols = list(keywords_map.keys())
        suma_clas = sum([resultado.get(col, 0) for col in clas_cols])
        resultado['Otros'] = resultado['cargo'] - suma_clas
        resultado['Redond'] = round(resultado['cargo'] - suma_clas - resultado['Otros'], 2)
        resultado['Saldo'] = 0.0

        return pd.Series(resultado)

    # Agrupar por asiento contable y aplicar clasificación
    df_grouped = df.groupby('Asiento contable').apply(clasificar_asiento).reset_index(drop=True)

    # Añadir índice inicial (como en archivo modelo)
    df_grouped.insert(0, 'Unnamed: 0', range(1, len(df_grouped) + 1))

    # Agregar fila de totales al final
    total_row = {col: df_grouped[col].sum() if df_grouped[col].dtype != 'O' else '' for col in df_grouped.columns}
    total_row['Concepto / Referencia'] = 'TOTAL'
    df_grouped.loc[len(df_grouped)] = total_row

    # Exportar a Excel
    df_grouped.to_excel(ruta_salida, index=False)

    # ========================
    # Estilización del Excel
    # ========================
    wb = load_workbook(ruta_salida)
    ws = wb.active

    # Fijar cabecera y activar autofiltro
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}1"

    # Ajuste automático de anchos de columna
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    # Definir colores y bordes para uso posterior
    fill_green = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    border = Border(bottom=Side(style="thin", color="000000"))

    # Resaltar celdas clasificadas con color verde
    clas_cols = list(keywords_map.keys())
    for col_idx, col_name in enumerate(df_grouped.columns):
        if col_name in clas_cols:
            col_num = col_idx + 1
            for row in range(2, ws.max_row):
                cell = ws.cell(row=row, column=col_num)
                if cell.value not in (None, 0, ''):
                    cell.fill = fill_green

    # Dibujar línea separadora entre fechas distintas
    last_date = None
    for i in range(2, ws.max_row):
        current_date = ws.cell(row=i, column=2).value
        if current_date != last_date and last_date is not None:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=i - 1, column=col).border = border
        last_date = current_date

    # Estilizar la fila TOTAL al final del archivo
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    wb.save(ruta_salida)
