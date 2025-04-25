# =============================
# Importación de librerías
# =============================

import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
import os
import subprocess
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
                                                                                                                                                    
# =============================
# Función de transformación principal
# =============================
def transformar_excel(ruta_entrada, ruta_salida):
    # Leer datos
    df = pd.read_excel(ruta_entrada)

    # Filtrar filas con Importe negativo
    df = df[~df['Importe'].astype(str).str.contains('-', na=False)]

    # Propagar Asiento contable y Fecha a filas de detalle
    df['Asiento contable'] = df['Asiento contable'].ffill()
    df['Fecha'] = df['Fecha'].ffill()

    # Auxiliares minúsculas para búsqueda
    df['linea'] = df['Líneas de factura'].fillna('').astype(str).str.lower()
    df['partner'] = df['Partner'].fillna('').astype(str).str.lower()

    # Mapeo de palabras clave a referencias
    
    keywords_map = {
        'Facta': ['facta'],
        'Master': ['max'],
        'Almacen': ['alm', 'almacen'],
        'Comision': ['comision', 'b/2025'],
        'Submarcell': ['submr'],
        'Linea 9': ['linea 9', 'línea 9'],
        'Caja de cobro': ['caja de cobro'],
        'Traspaso': ['traspaso'],
        'SAT': ['sat', 'servicio de administracion', 'servicios de administracion']
    }

    

    def clasificar_asiento(grupo):
        res = {}
        res['Día'] = pd.to_datetime(grupo['Fecha'].iloc[0]).date()
        res['Concepto / Referencia'] = grupo['Líneas de factura'].iloc[0]
        res['cargo'] = ""
        res['Abono'] = grupo['Importe'].sum()
        res['Referencia'] = grupo['Referencia'].iloc[0]
        # Clasificación
        for col, keys in keywords_map.items():
            mask = grupo.apply(lambda r: any(kw in r['linea'] for kw in keys)
                                or any(kw in r['partner'] for kw in keys), axis=1)
            if col == 'Traspaso':
                res[col] = grupo.loc[mask, 'Líneas de factura/Débito'].sum()
            else:
                res[col] = grupo.loc[mask, 'Líneas de factura/Débito'].sum() + grupo.loc[mask, 'Líneas de factura/Crédito'].sum()
     
        # Otros y Redond (resumen de montos no clasificados y ajuste por redondeo)
        # 1. suma_clas: suma de todos los importes asignados a categorías
        suma_clas = sum(res[c] for c in keywords_map)
        # 2. base: monto total de la transacción (Abono, ya que cargo se fija en 0)
        base = res['Abono']
        # 3. Redond: diferencia entre el monto total y la suma clasificada
        res['Redond'] = base - suma_clas 
        
        # Cálculo correcto de Saldo: Abono menos todas las categorías
        res['Saldo'] = res['Abono'] - sum(res[c] for c in keywords_map) - res['Redond']
        
        return pd.Series(res)

    # Agrupar y obtener DF final
    df_grouped = df.groupby('Asiento contable', dropna=False).apply(clasificar_asiento).reset_index(drop=True)
    df_grouped.insert(0, '#', range(1, len(df_grouped) + 1))
    # Totales
    total = {c: df_grouped[c].sum() if pd.api.types.is_numeric_dtype(df_grouped[c]) else ''
        for c in df_grouped.columns}
    total['Concepto / Referencia'] = 'TOTAL'
    df_grouped = pd.concat([df_grouped, pd.DataFrame([total])], ignore_index=True)

    # Exportar preliminar
    df_grouped.to_excel(ruta_salida, index=False)

    # ========================
    # Estilos y gráficos
    # ========================
    wb = load_workbook(ruta_salida)
    ws = wb.active

    # Cabecera
    header_fill = PatternFill('solid', start_color='4472C4', end_color='4472C4')
    header_font = Font(color='FFFFFF', bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Freeze y autofiltro
    ws.freeze_panes = 'B2'
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}1"
    
    # Auto-ajustar anchos de columna
    for col_cells in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col_cells if cell.value), default=0)
        ws.column_dimensions[col_cells[0].column_letter].width = max_length + 2

    # Colores por categoría
    palette = ['D9EAD3','FCE5CD','D0E0E3','EAD1DC','FFF2CC','C9DAF8','E2EFDA','E6B8B7',
            'FFD966','B6D7A8','EA9999','A4C2F4','D5A6BD','B7DEE8']
    color_map = {col: PatternFill('solid', start_color=palette[i % len(palette)],
                end_color=palette[i % len(palette)])
                for i, col in enumerate(keywords_map)}

    # Resaltar montos clasificados
    for idx, col in enumerate(df_grouped.columns, start=1):
        if col in color_map:
            for r in range(2, ws.max_row):
                cell = ws.cell(r, idx)
                if cell.value not in (None, 0, ''):
                    cell.fill = color_map[col]

    # Encontrar índices de columnas importantes
    concepto_ref_idx = None
    abono_idx = None
    redond_idx = None
    
    for idx, col in enumerate(df_grouped.columns, start=1):
        if col == 'Concepto / Referencia':
            concepto_ref_idx = idx
        elif col == 'Abono':
            abono_idx = idx
        elif col == 'Redond':
            redond_idx = idx
    
    # Colores personalizados
    deposito_fill = PatternFill('solid', start_color='FFCCCC', end_color='FFCCCC')  # Rosado
    traspaso_fill = PatternFill('solid', start_color='CCFFCC', end_color='CCFFCC')  # Verde
    abono_fill = PatternFill('solid', start_color='CCE5FF', end_color='CCE5FF')     # Azul
    
    # Aplicar formatos especiales
    for r in range(2, ws.max_row + 1):
        # Formatear números con separador de miles
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
        
        # Aplicar colores específicos
        if concepto_ref_idx:
            concepto_valor = str(ws.cell(r, concepto_ref_idx).value).lower()
            if 'deposito en efectivo' in concepto_valor:
                ws.cell(r, concepto_ref_idx).fill = deposito_fill
            elif 'traspaso' in concepto_valor:
                ws.cell(r, concepto_ref_idx).fill = traspaso_fill
        
        # Columna Abono en azul
        if abono_idx:
            ws.cell(r, abono_idx).fill = abono_fill
        
        # Redond negativo en rojo
        if redond_idx:
            redond_cell = ws.cell(r, redond_idx)
            if isinstance(redond_cell.value, (int, float)) and redond_cell.value < 0:
                redond_cell.font = Font(color='FF0000')
            if isinstance(redond_cell.value, (int, float)) and redond_cell.value > 0: redond_cell.font = Font(color='000000')
            if isinstance(redond_cell.value, (int, float)) and redond_cell.value == 0: redond_cell.font = Font(color='FFFFFF')
            
    
    # Sombreado alternado para filas (solo celdas sin estilo previo)
    alt_fill = PatternFill('solid', start_color='F2F2F2', end_color='F2F2F2')
    total_row = ws.max_row
    for r in range(2, total_row):
        if r % 2 == 0:
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(r, c)
                if cell.fill.fill_type is None:
                    cell.fill = alt_fill

    # Línea separadora entre fechas
    border = Border(bottom=Side(style='thin', color='000000'))
    last = None
    for r in range(2, ws.max_row):
        curr = ws.cell(r, 2).value
        if last and curr != last:
            for c in range(1, ws.max_column + 1):
                ws.cell(r - 1, c).border = border
        last = curr

    # Estilo TOTAL
    total_fill = PatternFill('solid', start_color='F4CCCC', end_color='F4CCCC')
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(ws.max_row, c)
        cell.fill = total_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

  

    wb.save(ruta_salida)

# =============================
# Interfaz gráfica Tkinter
# =============================
class App:
    def __init__(self, root):
        self.root = root
        self.root.title("📊 Transformador Excel Contable")
        self.root.geometry("700x420")
        self.root.configure(bg='#181a20')  # Fondo más oscuro
        self.ruta_entrada = ''
        self.ruta_salida = ''
        self.build_ui()

    def build_ui(self):
        # Título
        title = tk.Label(
            self.root,
            text="💼 Transformador Excel - Rosa Marcela",
            font=("Segoe UI", 22, "bold"),
            bg="#181a20",
            fg="#f1f2f6"
        )
        title.pack(pady=(18, 8))

        # Marco principal con borde y padding
        main_frame = tk.Frame(self.root, bg="#23272e", bd=0, relief="flat", highlightbackground="#444", highlightthickness=2)
        main_frame.pack(padx=28, pady=8, fill="both", expand=True)

        # Botones y etiquetas
        btn_frame = tk.Frame(main_frame, bg="#23272e")
        btn_frame.pack(pady=18)

        def on_enter(e): e.widget['bg'] = '#3b4252'
        def on_leave(e): e.widget['bg'] = e.widget.default_bg

        btn_cargar = tk.Button(
            btn_frame, text="📂 Cargar archivo Excel", font=("Segoe UI", 13, "bold"),
            bg="#0984e3", fg="#f1f2f6", activebackground="#74b9ff", activeforeground="#23272e",
            width=20, height=1, command=self.cargar_archivo, bd=0, relief="ridge", cursor="hand2"
        )
        btn_cargar.default_bg = "#0984e3"
        btn_cargar.bind("<Enter>", on_enter)
        btn_cargar.bind("<Leave>", on_leave)
        btn_cargar.grid(row=0, column=0, padx=10, pady=8, sticky="ew")

        btn_guardar = tk.Button(
            btn_frame, text="💾 Generar y guardar archivo", font=("Segoe UI", 13, "bold"),
            bg="#00b894", fg="#f1f2f6", activebackground="#55efc4", activeforeground="#23272e",
            width=24, height=1, command=self.guardar_archivo, bd=0, relief="ridge", cursor="hand2"
        )
        btn_guardar.default_bg = "#00b894"
        btn_guardar.bind("<Enter>", on_enter)
        btn_guardar.bind("<Leave>", on_leave)
        btn_guardar.grid(row=0, column=1, padx=10, pady=8, sticky="ew")

        self.btn_abrir = tk.Button(
            main_frame, text="📊 Abrir archivo generado", font=("Segoe UI", 13, "bold"),
            bg="#fdcb6e", fg="#23272e", activebackground="#ffeaa7", activeforeground="#23272e",
            width=28, height=1, command=self.abrir_archivo, bd=0, relief="ridge", cursor="hand2", state=tk.DISABLED
        )
        self.btn_abrir.default_bg = "#fdcb6e"
        self.btn_abrir.bind("<Enter>", on_enter)
        self.btn_abrir.bind("<Leave>", on_leave)
        self.btn_abrir.pack(side="bottom", pady=(22, 0), fill="x", padx=10)

        # Área de mensajes tipo consola
        msg_frame = tk.Frame(main_frame, bg="#181a20", bd=0, relief="flat")
        msg_frame.pack(fill="x", padx=18, pady=(8, 0))

        self.label_archivo = tk.Label(
            msg_frame, text="Archivo cargado: Ninguno",
            font=("Segoe UI", 10, "bold"), bg="#181a20", fg="#b2bec3", anchor="w", wraplength=600, pady=4
        )
        self.label_archivo.pack(fill="x")

        self.label_guardado = tk.Label(
            msg_frame, text="Archivo generado: Ninguno",
            font=("Segoe UI", 10, "bold"), bg="#181a20", fg="#b2bec3", anchor="w", wraplength=600, pady=4
        )
        self.label_guardado.pack(fill="x")

        # Pie de página
        footer = tk.Label(
            self.root, text="© 2025 Rosa Marcela", font=("Segoe UI", 9),
            bg="#181a20", fg="#636e72"
        )
        footer.pack(side="bottom", pady=5)

    def cargar_archivo(self):
        ruta = filedialog.askopenfilename(filetypes=[("Archivos Excel", "*.xlsx")])
        if ruta:
            self.ruta_entrada = ruta
            self.label_archivo.config(
                text=f"✅ Archivo cargado: {os.path.basename(ruta)}",
                fg="#00e676"  # Verde brillante
            )
            self.btn_abrir.config(state=tk.DISABLED)
        else:
            self.label_archivo.config(
                text="Archivo cargado: Ninguno",
                fg="#b2bec3"
            )
            self.btn_abrir.config(state=tk.DISABLED)

    def guardar_archivo(self):
        if not self.ruta_entrada:
            self.label_guardado.config(
                text="⚠️ Selecciona un archivo primero.",
                fg="#fdcb6e"  # Naranja
            )
            messagebox.showwarning("⚠️ Selecciona un archivo primero.")
            return
        ruta = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Archivos Excel","*.xlsx")])
        if ruta:
            try:
                transformar_excel(self.ruta_entrada, ruta)
                self.ruta_salida = ruta
                self.label_guardado.config(
                    text=f"✅ Generado: {os.path.basename(ruta)}",
                    fg="#00e676"  # Verde brillante
                )
                self.btn_abrir.config(state=tk.NORMAL)  # Habilita el botón
                messagebox.showinfo("✅ Éxito","Archivo creado y estilizado.")
            except Exception as e:
                self.label_guardado.config(
                    text=f"❌ Error: {str(e)}",
                    fg="#ff7675"  # Rojo claro
                )
                self.btn_abrir.config(state=tk.DISABLED)  # Deshabilita el botón en caso de error
                messagebox.showerror("❌ Error", str(e))

    def abrir_archivo(self):
        if self.ruta_salida and os.path.exists(self.ruta_salida):
            subprocess.Popen(['start','',self.ruta_salida], shell=True)
        else:
            messagebox.showwarning("⚠️","Archivo no encontrado.")

if __name__ == '__main__':
    root = tk.Tk(); App(root); root.mainloop()